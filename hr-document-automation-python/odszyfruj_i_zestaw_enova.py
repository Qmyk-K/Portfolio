# -*- coding: utf-8 -*-
r"""
odszyfruj_i_zestaw_enova_v16.py
Wersja: v16_zabezpieczenie_pola_poczta

Odszyfrowuje PDF-y "Oswiadczenie Zleceniobiorcy", wyciaga dane osobowe
i adresowe, dopasowuje kody Urzedu Skarbowego i NFZ ze slownikow,
waliduje kluczowe pola i zapisuje wynik do enova_import.xlsx gotowego
do importu w Enova365. Wszystko niepewne trafia do arkusza KONFLIKTY.

INSTALACJA (jednorazowo, w PowerShell):
    pip install pypdf cryptography openpyxl

UWAGA: pakiet "cryptography" jest wymagany przez pypdf do odszyfrowania PDF-ow
zabezpieczonych algorytmem AES (typowe dla PDF-ow eksportowanych z Adobe/Worda
z haslem). Bez niego decrypt() konczy sie wyjatkiem dla KAZDEGO pliku, co
objawia sie pustym folderem _Odszyfrowane_PDF i brakiem enova_import.xlsx -
dokladnie tak jak przy pierwszym uruchomieniu bez tej biblioteki.

STRUKTURA FOLDEROW (obok tego skryptu):
    PDF\                    <- zaszyfrowane PDF-y (czytane tez z podfolderow)
    Slowniki\               <- US.csv i NFZ.csv (UTF-16, tab-separated, kolumny Kod/Nazwa)
    enova_import.xlsx       <- plik wynikowy (tworzony przez skrypt)
    _Odszyfrowane_PDF_RRRR_MM_DD-GG_MM\   <- jawne kopie PDF

UWAGA (do zweryfikowania przez operatora):
    Nazwy kolumn "zawsze pustych" (telefony, faks, skrzynka pocztowa,
    dokument tozsamosci) sa nazwami roboczymi wg konwencji Enova,
    NIE zostaly potwierdzone w zadnym realnym szablonie importu.
    Przed pierwszym importem porownaj liste ENOVA_COLUMNS z prawdziwym
    szablonem importu Enova i popraw nazwy, ktore sie nie zgadzaja.

URUCHOMIENIE:
    python odszyfruj_i_zestaw_enova_v16.py
"""

import os
import re
import sys
import glob
import difflib
import unicodedata
from datetime import datetime

from pypdf import PdfReader, PdfWriter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def pokaz_komunikat_windows(tytul, tresc, ikona="info"):
    """Pokazuje natywne okienko Windows (MessageBox), zeby wazny komunikat
    (np. rozne firmy w jednym imporcie, blad krytyczny) nie zniknal razem
    z zamknieciem konsoli. Poza Windows tylko drukuje do konsoli."""
    print(f"\n[{tytul}] {tresc}")
    if os.name != "nt":
        return
    try:
        import ctypes
        ikony = {"info": 0x40, "warning": 0x30, "error": 0x10}
        # MB_OK (0x0) + MB_SETFOREGROUND (0x10000), zeby okno wyskoczylo na wierzch
        flagi = 0x0 | 0x00010000 | ikony.get(ikona, 0x40)
        ctypes.windll.user32.MessageBoxW(0, tresc, tytul, flagi)
    except Exception:
        pass  # brak GUI/ctypes nie moze wywalic calego skryptu

# ======================= KONFIGURACJA =======================
SKRYPT_DIR = os.path.dirname(os.path.abspath(__file__))

HASLO = "TWOJE_HASLO_TUTAJ"
FOLDER_PDF = os.path.join(SKRYPT_DIR, "PDF")
FOLDER_SLOWNIKI = os.path.join(SKRYPT_DIR, "Slowniki")
PLIK_WYJSCIOWY_XLSX = os.path.join(SKRYPT_DIR, "enova_import.xlsx")

_teraz = datetime.now()
FOLDER_ODSZYFROWANE = os.path.join(
    SKRYPT_DIR,
    f"_Odszyfrowane_PDF_{_teraz.strftime('%Y_%m_%d-%H_%M')}",
)

PROG_FUZZY = 0.86
UZUPELNIJ_KRAJ_POLSKA = True

# Kolor podswietlenia duplikatow (PESEL)
KOLOR_DUPLIKAT = "FF39FDB2"

# Mapowanie wojewodztwo -> kod regionalny NFZ (fallback status REGION)
NFZ_REGION_KOD = {
    "dolnoslaskie": "01R",
    "kujawsko-pomorskie": "02R",
    "lubelskie": "03R",
    "lubuskie": "04R",
    "lodzkie": "05R",
    "malopolskie": "06R",
    "mazowieckie": "07R",
    "opolskie": "08R",
    "podkarpackie": "09R",
    "podlaskie": "10R",
    "pomorskie": "11R",
    "slaskie": "12R",
    "swietokrzyskie": "13R",
    "warminsko-mazurskie": "14R",
    "wielkopolskie": "15R",
    "zachodniopomorskie": "16R",
}

# ======================= KOLUMNY ENOVA =======================
# Dokladna lista i kolejnosc kolumn potwierdzona przez Kamila (realny szablon importu).
# UWAGA: Last.Adres ma Gmine na poczatku bloku (za Wojewodztwem), a Powiat na koncu bloku.
#        Last.AdresZamieszkania / Last.AdresDoKorespondencji NIE MAJA pola Gmina,
#        a Powiat jest zaraz po Wojewodztwie (inny uklad niz w Last.Adres).
#        Telefon/Faks sa pod Last.Adres (nie pod Kontakt), a TelefonKomorkowy
#        i SkrytkaPocztowa sa pod Last.Kontakt. Wszystkie te pola formularz
#        nie dostarcza, wiec zostaja zawsze puste.
ENOVA_COLUMNS = [
    "NazwaPliku", "Class", "Kod",
    "Last.Nazwisko", "Last.Imie", "Last.ImieDrugie",
    "Last.ImieOjca", "Last.ImieMatki",
    "Last.NazwiskoRodowe", "Last.NIP",
    "Last.Dokument.Rodzaj", "Last.Dokument.SeriaNumer",
    "Last.PESEL", "Last.Obywatelstwo.Nazwa",
    "Last.Urodzony.Data", "Last.Urodzony.Miejsce",
    "Last.Adres.Kraj", "Last.Adres.Wojewodztwo", "Last.Adres.Gmina",
    "Last.Adres.Ulica", "Last.Adres.NrDomu", "Last.Adres.NrLokalu",
    "Last.Adres.Miejscowosc", "Last.Adres.KodPocztowyS", "Last.Adres.Poczta",
    "Last.Adres.Powiat",
    "Kontakt.EMAIL", "Last.Kontakt.TelefonKomorkowy", "Last.Kontakt.SkrytkaPocztowa",
    "Last.Adres.Telefon", "Last.Adres.Faks",
    "Last.AdresZamieszkania.Kraj", "Last.AdresZamieszkania.Wojewodztwo",
    "Last.AdresZamieszkania.Powiat", "Last.AdresZamieszkania.Ulica",
    "Last.AdresZamieszkania.NrDomu", "Last.AdresZamieszkania.NrLokalu",
    "Last.AdresZamieszkania.Miejscowosc", "Last.AdresZamieszkania.KodPocztowy",
    "Last.AdresZamieszkania.Poczta",
    "Last.AdresDoKorespondencji.Kraj", "Last.AdresDoKorespondencji.Wojewodztwo",
    "Last.AdresDoKorespondencji.Powiat", "Last.AdresDoKorespondencji.Ulica",
    "Last.AdresDoKorespondencji.NrDomu", "Last.AdresDoKorespondencji.NrLokalu",
    "Last.AdresDoKorespondencji.Miejscowosc", "Last.AdresDoKorespondencji.KodPocztowy",
    "Last.AdresDoKorespondencji.Poczta",
    "Last.StNiezdolnDoPracy.Kod", "Last.OddzialNFZ.Kod", "Last.OddzialNFZ.OdDnia",
    "Last.Podatki.UrzadSkarbowy:Kod",
    "Rachunki:Class", "Rachunki:Rachunek.Numer.CS",
    "Rachunki:Rachunek.Numer.Kierunek", "Rachunki:Rachunek.Numer.Numer",
]

# (prefix, ma_gmine, sufiks_kodu_pocztowego) dla trzech blokow adresowych
ADRESY_SPEC = [
    ("Last.Adres", True, "S"),
    ("Last.AdresZamieszkania", False, ""),
    ("Last.AdresDoKorespondencji", False, ""),
]


def wypelnij_adres(data, prefix, addr, ma_gmine, sufiks_kodu):
    data[f"{prefix}.Wojewodztwo"] = addr["wojewodztwo"]
    data[f"{prefix}.Powiat"] = addr["powiat"]
    if ma_gmine:
        data[f"{prefix}.Gmina"] = addr["gmina"]
    data[f"{prefix}.Ulica"] = addr["ulica"]
    data[f"{prefix}.NrDomu"] = addr["nr_domu"]
    data[f"{prefix}.NrLokalu"] = addr["nr_mieszkania"]
    data[f"{prefix}.Miejscowosc"] = addr["miejscowosc"]
    data[f"{prefix}.KodPocztowy{sufiks_kodu}"] = addr["kod_pocztowy"]
    data[f"{prefix}.Poczta"] = addr["poczta"]
    if UZUPELNIJ_KRAJ_POLSKA and any(addr.values()):
        data[f"{prefix}.Kraj"] = "Polska"

KONFLIKTY_COLUMNS = [
    "Plik", "PESEL", "Nazwisko", "Imie", "Priorytet", "Kategoria",
    "Problem", "Sugerowana_akcja", "Pola_roznic",
]


# ======================= NORMALIZACJA =======================

def clean(s):
    if s is None:
        return ""
    s = s.strip()
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"\s+([,.;:])", r"\1", s)
    s = re.sub(r"[:;,.\u2014\-]+$", "", s).strip()
    return s


def _strip_diacritics(s):
    return "".join(
        c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c)
    )


def normalize_for_match(s):
    s = _strip_diacritics(s or "").lower()
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def only_digits(s):
    return re.sub(r"\D", "", s or "")


_RZYMSKIE = re.compile(r"^(M{0,4}(CM|CD|D?C{0,3})(XC|XL|L?X{0,3})(IX|IV|V?I{0,3}))$", re.IGNORECASE)


def to_title(s):
    if not s:
        return ""
    parts = re.split(r"([ \-])", s.lower())
    out = []
    for p in parts:
        if p in (" ", "-") or p == "":
            out.append(p)
        else:
            out.append(p[:1].upper() + p[1:])
    return "".join(out)


def to_title_street(s):
    if not s:
        return ""
    tokens = s.split(" ")
    out = []
    for tok in tokens:
        core = tok.strip(".")
        if core and _RZYMSKIE.match(core) and core.upper() == core:
            out.append(tok.upper())
        else:
            out.append(to_title(tok))
    return " ".join(out)


def skroc_kod_pocztowy(kod):
    """05-500 -> 5500 (Enova wymaga bez wiodacego zera). ZAMIERZONE."""
    cyfry = only_digits(kod)
    if not cyfry:
        return ""
    cyfry = cyfry.lstrip("0")
    return cyfry if cyfry else "0"


def waliduj_pesel(pesel):
    if not re.fullmatch(r"\d{11}", pesel or ""):
        return False
    wagi = [1, 3, 7, 9, 1, 3, 7, 9, 1, 3]
    suma = sum(int(pesel[i]) * wagi[i] for i in range(10))
    kontrolna = (10 - (suma % 10)) % 10
    return kontrolna == int(pesel[10])


def waliduj_nip(nip):
    cyfry = only_digits(nip)
    if len(cyfry) != 10:
        return False
    wagi = [6, 5, 7, 2, 3, 4, 5, 6, 7]
    suma = sum(int(cyfry[i]) * wagi[i] for i in range(9))
    kontrolna = suma % 11
    if kontrolna == 10:
        return False
    return kontrolna == int(cyfry[9])


def waliduj_email(email):
    return bool(re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", email or ""))


# ======================= SLOWNIKI (US / NFZ) =======================

def wczytaj_slownik(path):
    """Wczytuje CSV UTF-16, tab-separated, kolumny Kod/Nazwa (+ opcjonalnie Adres).
    Zwraca liste (kod, nazwa, adres). adres = "" jesli kolumny nie ma w pliku."""
    if not os.path.isfile(path):
        return None
    wpisy = []
    with open(path, "r", encoding="utf-16") as f:
        naglowek = f.readline()
        kolumny = [c.strip() for c in naglowek.split("\t")]
        try:
            idx_kod = kolumny.index("Kod")
            idx_nazwa = kolumny.index("Nazwa")
        except ValueError:
            idx_kod, idx_nazwa = 0, 1
        idx_adres = kolumny.index("Adres") if "Adres" in kolumny else None
        for linia in f:
            linia = linia.rstrip("\n\r")
            if not linia.strip():
                continue
            czesci = linia.split("\t")
            if len(czesci) <= max(idx_kod, idx_nazwa):
                continue
            adres = ""
            if idx_adres is not None and len(czesci) > idx_adres:
                adres = czesci[idx_adres].strip()
            wpisy.append((czesci[idx_kod].strip(), czesci[idx_nazwa].strip(), adres))
    return wpisy


_KOD_POCZTOWY_RE = re.compile(r"\d{2}-\d{3}")


def dopasuj_urzad(nazwa_z_pdf, slownik, prog_fuzzy=PROG_FUZZY):
    """Zwraca (kod, status). Status: DOKLADNE, ZAWIERA, PRZYBLIZONE,
    NIEJEDNOZNACZNE, BRAK_DOPASOWANIA, BRAK, BRAK_SLOWNIKA.

    Gdy dopasowanie nazwy daje kilka trafien (typowo: PDF pisze nazwe urzedu
    bez liczebnika porzadkowego - "Urzad Skarbowy Warszawa - Srodmiescie" -
    a slownik ma osobne wpisy "Pierwszy/Drugi/Trzeci ..."), probujemy
    rozstrzygnac po kodzie pocztowym z linii PDF (jest w niej zawsze,
    bo to ta sama linia co adres urzedu) porownujac go z kodem pocztowym
    w kolumnie Adres slownika. To dziala nawet gdy nazwy sa identyczne."""
    if not nazwa_z_pdf or not nazwa_z_pdf.strip():
        return "", "BRAK"
    if slownik is None:
        return "", "BRAK_SLOWNIKA"

    cel = normalize_for_match(nazwa_z_pdf)
    if not cel:
        return "", "BRAK"

    kod_pocztowy_pdf = None
    m_kod = _KOD_POCZTOWY_RE.search(nazwa_z_pdf)
    if m_kod:
        kod_pocztowy_pdf = m_kod.group(0)

    def zawez_po_kodzie_pocztowym(kandydaci):
        """kandydaci: lista (kod, nazwa, adres). Zwraca liste kodow po zawezeniu."""
        if not kod_pocztowy_pdf:
            return None
        zawezone = [k for k, n, a in kandydaci if kod_pocztowy_pdf in a]
        zawezone = list(dict.fromkeys(zawezone))
        return zawezone

    # 1. dokladne
    trafienia_pelne = [(k, n, a) for k, n, a in slownik if normalize_for_match(n) == cel]
    if len(trafienia_pelne) == 1:
        return trafienia_pelne[0][0], "DOKLADNE"
    if len(trafienia_pelne) > 1:
        zawezone = zawez_po_kodzie_pocztowym(trafienia_pelne)
        if zawezone and len(zawezone) == 1:
            return zawezone[0], "DOKLADNE"
        return "", "NIEJEDNOZNACZNE"

    # 2. zawiera (jedna nazwa zawiera druga)
    trafienia_pelne = []
    for k, n, a in slownik:
        nn = normalize_for_match(n)
        if not nn:
            continue
        if nn in cel or cel in nn:
            trafienia_pelne.append((k, n, a))
    # usun duplikaty po kodzie
    widziane = set()
    unikalne = []
    for rec in trafienia_pelne:
        if rec[0] not in widziane:
            widziane.add(rec[0])
            unikalne.append(rec)
    trafienia_pelne = unikalne

    if len(trafienia_pelne) == 1:
        return trafienia_pelne[0][0], "ZAWIERA"
    if len(trafienia_pelne) > 1:
        zawezone = zawez_po_kodzie_pocztowym(trafienia_pelne)
        if zawezone and len(zawezone) == 1:
            return zawezone[0], "ZAWIERA"
        return "", "NIEJEDNOZNACZNE"

    # 3. przyblizone (fuzzy)
    wyniki = []
    for k, n, a in slownik:
        wynik = difflib.SequenceMatcher(None, cel, normalize_for_match(n)).ratio()
        wyniki.append((wynik, k, n, a))
    wyniki.sort(key=lambda x: x[0], reverse=True)

    if wyniki and wyniki[0][0] >= prog_fuzzy:
        najlepszy_wynik = wyniki[0][0]
        # kandydaci blisko najlepszego wyniku (roznica < 0.02) traktujemy jako remis
        remisowi = [w for w in wyniki if najlepszy_wynik - w[0] < 0.02]
        if len(remisowi) == 1:
            return remisowi[0][1], "PRZYBLIZONE"
        zawezone = zawez_po_kodzie_pocztowym([(w[1], w[2], w[3]) for w in remisowi])
        if zawezone and len(zawezone) == 1:
            return zawezone[0], "PRZYBLIZONE"
        return "", "NIEJEDNOZNACZNE"

    return "", "BRAK_DOPASOWANIA"


def dopasuj_nfz_region(wojewodztwo):
    klucz_bez_myslnika = normalize_for_match(wojewodztwo)
    for k, kod in NFZ_REGION_KOD.items():
        if normalize_for_match(k) == klucz_bez_myslnika:
            return kod
    return ""


def dopasuj_nfz_po_pierwszym_slowie(nazwa_z_pdf, slownik):
    """Realne wpisy NFZ.csv maja format '<Przymiotnik> Oddzial Wojewodzki NFZ w <Miasto>',
    a PDF zwykle pisze '<Przymiotnik> Oddzial Narodowego Funduszu Zdrowia w <Miasto>'.
    Reszta zdania sie rozjezdza (skrot NFZ vs pelna nazwa, inne miasto w adresie),
    wiec dopasowanie po calym stringu czesto zawodzi. Pierwsze slowo (przymiotnik
    wojewodztwa) jest za to stabilne i praktycznie unikalne w slowniku."""
    if slownik is None:
        return "", "BRAK_SLOWNIKA"
    pierwsze_slowo = normalize_for_match(nazwa_z_pdf).split(" ")[0] if nazwa_z_pdf else ""
    if not pierwsze_slowo:
        return "", "BRAK"
    trafienia = []
    for kod, nazwa, _adres in slownik:
        pierwsze_slownik = normalize_for_match(nazwa).split(" ")[0] if nazwa else ""
        if pierwsze_slownik and pierwsze_slownik == pierwsze_slowo:
            trafienia.append(kod)
    trafienia = list(dict.fromkeys(trafienia))
    if len(trafienia) == 1:
        return trafienia[0], "DOKLADNE"
    if len(trafienia) > 1:
        return "", "NIEJEDNOZNACZNE"
    return "", "BRAK_DOPASOWANIA"


# ======================= EKSTRAKCJA TEKSTU PDF =======================

def wyciagnij_tekst(reader, tryb):
    strony = []
    for page in reader.pages:
        if tryb == "layout":
            try:
                t = page.extract_text(extraction_mode="layout") or ""
            except TypeError:
                t = page.extract_text() or ""
        else:
            t = page.extract_text() or ""
        strony.append(t)
    return "\n".join(strony)


# ======================= PARSOWANIE ADRESU =======================

def parse_address_block_lines(lines):
    """9 pol: wojewodztwo, powiat, gmina, ulica, nr_domu, nr_mieszkania,
    miejscowosc, kod_pocztowy, poczta. 'Taki sam jak' -> wszystkie puste.

    Pole 'poczta' to w praktyce kosz, do ktorego ludzie potrafia wpisac
    cokolwiek (ulice, drugi kod pocztowy, adres email). Skrypt i tak przepisuje
    to co znalazl do kolumny Poczta (operator ma to widziec w Enovie), ale
    gdy wartosc nie wyglada jak czysta nazwa miejscowosci (zawiera cyfry albo
    '@'), dodatkowo ustawia flage 'poczta_niepewna', zeby parse_pdf_text mogl
    dodac konflikt do recznego sprawdzenia."""
    result = {
        "wojewodztwo": "", "powiat": "", "gmina": "",
        "ulica": "", "nr_domu": "", "nr_mieszkania": "",
        "miejscowosc": "", "kod_pocztowy": "", "poczta": "",
        "poczta_niepewna": False,
    }
    full_text = " ".join(lines)
    if "taki sam jak" in full_text.lower():
        return result  # celowo puste - operator wie ze to ten sam adres

    for line in lines:
        line = line.strip()
        if not line:
            continue

        m = re.match(
            r"Wojew[oó]dztwo\s+([^,]+),\s*powiat\s+([^,]+),\s*gmina:?\s*(.+)",
            line, re.IGNORECASE,
        )
        if m:
            result["wojewodztwo"] = clean(m.group(1)).lower()
            result["powiat"] = clean(m.group(2))
            result["gmina"] = to_title(clean(m.group(3)))
            continue

        m = re.match(r"^(\d{2}-\d{3})\s+(.+)$", line)
        if m:
            result["kod_pocztowy"] = skroc_kod_pocztowy(m.group(1))
            poczta_raw = clean(m.group(2))
            result["poczta"] = to_title(poczta_raw)
            if poczta_raw and re.search(r"[\d@]", poczta_raw):
                result["poczta_niepewna"] = True
            continue

        if "," in line:
            street_part, city_part = line.rsplit(",", 1)
            result["miejscowosc"] = to_title(clean(city_part))
            street_part = clean(street_part)
            street_part = re.sub(r"^ul\.\s*", "", street_part, flags=re.IGNORECASE)
            m2 = re.match(r"^(.*?)\s+(\d+[A-Za-z]?)(?:/(\S+))?$", street_part)
            if m2:
                result["ulica"] = to_title_street(m2.group(1).strip())
                result["nr_domu"] = m2.group(2).strip()
                result["nr_mieszkania"] = (m2.group(3) or "").strip()
            else:
                result["ulica"] = to_title_street(street_part)
            continue

    return result


def split_address_blocks(text_default):
    """Dzieli tekst na 3 bloki adresowe wg etykiet w osobnych liniach."""
    def wytnij(poczatek, konce):
        wzorzec_start = re.escape(poczatek)
        for koniec in konce:
            m = re.search(
                wzorzec_start + r"\s*\n(.*?)\n\s*" + re.escape(koniec),
                text_default, re.DOTALL,
            )
            if m:
                return [l.strip() for l in m.group(1).split("\n") if l.strip()]
        return []

    zameld = wytnij("Adres zameldowania", ["Adres zamieszkania"])
    zamieszkania = wytnij("Adres zamieszkania", ["Adres do korespondencji"])
    korespondencji = wytnij(
        "Adres do korespondencji",
        ["Urząd Skarbowy", "Urzad Skarbowy", "Właściwy Oddział", "Wlasciwy Oddzial", "Sposób wypłaty", "Sposob wyplaty"],
    )
    return zameld, zamieszkania, korespondencji


# ======================= PARSOWANIE CALEGO PDF =======================

RACHUNEK_MIN_CYFR = 20
RACHUNEK_DLUGOSC_NRB = 26


RACHUNKI_CLASS = "RachunekBankowyPracownika"


def podziel_rachunek(cyfry):
    """Dzieli cyfry NRB na (cyfra_kontrolna_CS, kierunek_8_cyfr, numer_reszta).
    Zachowuje wiodace zera w kazdym segmencie (to pola tekstowe w Enova)."""
    cs = cyfry[:2]
    kierunek = cyfry[2:10]
    numer = cyfry[10:]
    return cs, kierunek, numer


def wyciagnij_rachunek(text, conflicts):
    """Zwraca (cs, kierunek, numer) dla dokladnie jednego wiarygodnego numeru
    rachunku, albo ("", "", "") gdy brak numeru (gotowka) lub gdy sytuacja
    jest niejednoznaczna (kilka numerow w PDF - wtedy tylko konflikt,
    zeby nie wpisac do Enovy przypadkowo zlego rachunku)."""
    m = re.search(
        r"Numer rachunku bankowego[^\n]*:\s*\n(.*?)(?:\nAdres zameldowania|\Z)",
        text, re.DOTALL | re.IGNORECASE,
    )
    if not m:
        return "", "", ""

    fragment = m.group(1)
    if re.search(r"do r[ąa]k w[łl]asnych|got[oó]wk", fragment, re.IGNORECASE):
        return "", "", ""  # wyplata gotowkowa, brak konfliktu

    ciagi = re.findall(r"[\d\s]{20,}", fragment)
    numery = []
    for c in ciagi:
        cyfry = only_digits(c)
        if len(cyfry) >= RACHUNEK_MIN_CYFR:
            numery.append(cyfry)

    if not numery:
        return "", "", ""

    if len(numery) > 1:
        conflicts.append(("Rachunek bankowy", "WYSOKI",
                           f"Znaleziono kilka numerow rachunku w jednym PDF ({', '.join(numery)}) - "
                           f"kolumny Rachunki zostawione puste, uzupelnij recznie"))
        return "", "", ""

    cyfry = numery[0]
    if len(cyfry) != RACHUNEK_DLUGOSC_NRB:
        conflicts.append(("Rachunek bankowy", "SREDNI",
                           f"Numer ma {len(cyfry)} cyfr, oczekiwano {RACHUNEK_DLUGOSC_NRB} - sprawdz podzial na CS/Kierunek/Numer"))

    return podziel_rachunek(cyfry)


def parse_pdf_text(text, filename):
    data = {kol: "" for kol in ENOVA_COLUMNS}
    conflicts = []  # (kategoria, priorytet, problem)

    data["NazwaPliku"] = clean(os.path.splitext(filename)[0])
    data["Class"] = "PracownikFirmy"

    # --- dane osobowe ---
    m = re.search(r"Nazwisko:\s*(.+?)\s+Imiona:[ \t]*(.+)", text)
    nazwisko = clean(m.group(1)) if m else ""
    imiona = clean(m.group(2)).split() if m else []
    data["Last.Nazwisko"] = to_title(nazwisko)
    data["Last.Imie"] = to_title(imiona[0]) if len(imiona) > 0 else ""
    data["Last.ImieDrugie"] = to_title(imiona[1]) if len(imiona) > 1 else ""
    if not nazwisko or not data["Last.Imie"]:
        conflicts.append(("Dane osobowe", "WYSOKI", "Brak Nazwiska lub Imienia"))

    m = re.search(r"Nazwisko rodowe:\s*(.+?)\s+Kraj obywatelstwa:", text)
    data["Last.NazwiskoRodowe"] = to_title(clean(m.group(1))) if m else ""

    # Imie ojca / Imie matki - wystepuje tylko w niektorych wariantach formularza
    m = re.search(r"Imi[eę] ojca:\s*(.+?)\s+Imi[eę] matki:[ \t]*(.+)", text)
    data["Last.ImieOjca"] = to_title(clean(m.group(1))) if m else ""
    data["Last.ImieMatki"] = to_title(clean(m.group(2))) if m else ""

    m = re.search(r"Data urodzenia:[ \t]*([\d.]+)", text)
    data["Last.Urodzony.Data"] = clean(m.group(1)) if m else ""

    m = re.search(r"Miejsce urodzenia:[ \t]*(.+)", text)
    data["Last.Urodzony.Miejsce"] = to_title(clean(m.group(1))) if m else ""

    m = re.search(r"Kraj obywatelstwa:[ \t]*(.+)", text)
    obywatelstwo_deklarowane = clean(m.group(1)) if m else ""
    if normalize_for_match(obywatelstwo_deklarowane) in ("polskie", "polska", "rp", "polish"):
        data["Last.Obywatelstwo.Nazwa"] = "polskie"
    else:
        data["Last.Obywatelstwo.Nazwa"] = ""
        if obywatelstwo_deklarowane:
            conflicts.append(("Obywatelstwo", "SREDNI",
                               f"Obywatelstwo zadeklarowane w PDF inne niz polskie ({obywatelstwo_deklarowane}) "
                               f"- uzupelnij Last.Obywatelstwo.Nazwa recznie"))
        else:
            conflicts.append(("Obywatelstwo", "SREDNI",
                               "Nie wykryto Kraju obywatelstwa w PDF - uzupelnij Last.Obywatelstwo.Nazwa recznie"))

    # PESEL / NIP tylko z bloku danych osobowych, NIE ze stopki RODO.
    # Niektore warianty formularza (np. aneks do umowy o dzielo) nie maja
    # w ogole "NIP:" na tej samej linii co PESEL - wtedy proba z NIP zawodzi
    # i lapiemy PESEL osobnym, prostszym wzorcem.
    m = re.search(r"PESEL:[ \t]*(\S+)\s+NIP:[ \t]*(\S+)", text)
    if m:
        pesel = only_digits(m.group(1))
        nip_raw = clean(m.group(2))
    else:
        m = re.search(r"PESEL:[ \t]*(\S+)", text)
        pesel = only_digits(m.group(1)) if m else ""
        nip_raw = ""
    data["Last.PESEL"] = pesel
    if pesel and not waliduj_pesel(pesel):
        conflicts.append(("PESEL", "WYSOKI", "PESEL nie przechodzi walidacji sumy kontrolnej lub zla dlugosc"))

    if normalize_for_match(nip_raw) in ("brak", ""):
        data["Last.NIP"] = ""
    else:
        nip_cyfry = only_digits(nip_raw)
        data["Last.NIP"] = nip_cyfry
        if nip_cyfry and not waliduj_nip(nip_cyfry):
            conflicts.append(("NIP", "SREDNI", "NIP podany, ale nie przechodzi walidacji sumy kontrolnej"))

    # Dokument tozsamosci - na razie obslugujemy TYLKO paszport (wg instrukcji).
    # Jesli przed numerem jest slowo "paszport" (w dowolnej odmianie: paszportu,
    # paszportem...) zakonczone dwukropkiem, wpisujemy Last.Dokument.Rodzaj =
    # "Paszport" i bierzemy numer po dwukropku. W kazdym innym przypadku (inny
    # rodzaj dokumentu albo brak wzmianki o paszporcie) OBA pola zostaja puste.
    m = re.search(r"[Pp]aszport\w*\s*:[ \t]*(\S+)", text)
    if m:
        data["Last.Dokument.Rodzaj"] = "Paszport"
        data["Last.Dokument.SeriaNumer"] = clean(m.group(1))
    else:
        data["Last.Dokument.Rodzaj"] = ""
        data["Last.Dokument.SeriaNumer"] = ""
        if not pesel:
            conflicts.append(("Dane osobowe", "SREDNI",
                               "Brak PESEL i brak wykrytego numeru paszportu - sprawdz PDF recznie "
                               "(inne dokumenty tozsamosci niz paszport nie sa jeszcze obslugiwane)"))

    # --- ZLECENIODAWCA / ZAMAWIAJACY (firma) ---
    # "Zamawiajacy" to etykieta uzywana w aneksach do umowy o dzielo,
    # "Zleceniodawca" w oswiadczeniach do umowy zlecenia.
    m = re.search(r"(?:ZLECENIODAWCA|ZAMAWIAJ[ĄA]CY):\s*(.+?),\s*NIP:[ \t]*(\d+)", text)
    if m:
        firma_nazwa = clean(m.group(1))
        firma_nip = clean(m.group(2))
    else:
        m = re.search(r"(?:ZLECENIODAWCA|ZAMAWIAJ[ĄA]CY):[ \t]*(.+)", text)
        firma_nazwa = clean(m.group(1)).rstrip(",").strip() if m else ""
        firma_nip = ""

    # --- email (sekcja zgody) ---
    m = re.search(r"adres email:[ \t]*\n(.*)", text)
    email = clean(m.group(1)) if m else ""
    if email:
        if waliduj_email(email):
            data["Kontakt.EMAIL"] = email
        else:
            data["Kontakt.EMAIL"] = email
            conflicts.append(("E-mail", "SREDNI", f"Email nie przechodzi walidacji: {email}"))

    # --- rachunek bankowy ---
    cs, kierunek, numer = wyciagnij_rachunek(text, conflicts)
    if numer:  # numer glowny (reszta cyfr) jest najlepszym wskaznikiem "cos znalezlismy"
        data["Rachunki:Class"] = RACHUNKI_CLASS
        data["Rachunki:Rachunek.Numer.CS"] = cs
        data["Rachunki:Rachunek.Numer.Kierunek"] = kierunek
        data["Rachunki:Rachunek.Numer.Numer"] = numer

    # --- adresy ---
    zameld_lines, zam_lines, kor_lines = split_address_blocks(text)
    zameld = parse_address_block_lines(zameld_lines)
    zamieszkania = parse_address_block_lines(zam_lines)
    korespondencji = parse_address_block_lines(kor_lines)

    for (prefix, ma_gmine, sufiks_kodu), addr in zip(
        ADRESY_SPEC, [zameld, zamieszkania, korespondencji]
    ):
        # Poczta zawiera cos innego niz czysta nazwa miejscowosci (cyfry, email,
        # ulica dopisana przez zleceniobiorce itp.) - wartosc i tak trafia do
        # kolumny (operator ma to zobaczyc w Enovie), konflikt tylko ostrzega.
        if addr["poczta_niepewna"]:
            conflicts.append(("Adres - Poczta", "WYSOKI",
                               f"Pole Poczta ({prefix}) zawiera cos innego niz sama nazwa miejscowosci "
                               f"({addr['poczta']!r}) - sprawdz i popraw recznie"))
        elif not addr["poczta"] and addr["kod_pocztowy"]:
            conflicts.append(("Adres - Poczta", "SREDNI",
                               f"Jest kod pocztowy ({prefix}) ale brak nazwy miejscowosci pocztowej - uzupelnij recznie"))

        wypelnij_adres(data, prefix, addr, ma_gmine, sufiks_kodu)

    # --- Urzad Skarbowy ---
    m = re.search(r"Urz[ąa]d Skarbowy \(.*?\):[ \t]*\n(.+)", text)
    us_nazwa = clean(m.group(1)) if m else ""
    slownik_us = wczytaj_slownik(os.path.join(FOLDER_SLOWNIKI, "US.csv"))
    kod_us, status_us = dopasuj_urzad(us_nazwa, slownik_us)
    data["Last.Podatki.UrzadSkarbowy:Kod"] = kod_us
    if status_us in ("BRAK_SLOWNIKA",):
        conflicts.append(("Slownik US", "WYSOKI", "Brak pliku US.csv w Slowniki"))
    elif status_us not in ("DOKLADNE", "ZAWIERA", "PRZYBLIZONE"):
        conflicts.append(("Urzad Skarbowy", "WYSOKI", f"Nie dopasowano ({status_us}): {us_nazwa}"))

    # --- NFZ ---
    # Umowa o dzielo (w odroznieniu od zlecenia) nie wymaga ubezpieczenia
    # zdrowotnego, wiec sekcja NFZ moze w ogole nie wystepowac w dokumencie.
    # Zgadywanie kodu NFZ po wojewodztwie ma sens TYLKO gdy etykieta NFZ
    # faktycznie byla w PDF (a jej wartosc po prostu nie dopasowala sie do
    # slownika) - jesli etykiety brak, NFZ zostaje pusty bez zadnego konfliktu.
    m = re.search(r"Oddzia[łl] Narodowego Funduszu Zdrowia:?[ \t]*\n(.+)", text)
    if not m:
        kod_nfz, status_nfz = "", "BRAK"
    else:
        nfz_nazwa = clean(m.group(1))
        slownik_nfz = wczytaj_slownik(os.path.join(FOLDER_SLOWNIKI, "NFZ.csv"))
        kod_nfz, status_nfz = dopasuj_nfz_po_pierwszym_slowie(nfz_nazwa, slownik_nfz)
        if status_nfz not in ("DOKLADNE",):
            # falbacki: dopasowanie ogolne (dokladne/zawiera/przyblizone), potem region z wojewodztwa
            kod_alt, status_alt = dopasuj_urzad(nfz_nazwa, slownik_nfz)
            if status_alt in ("DOKLADNE", "ZAWIERA", "PRZYBLIZONE"):
                kod_nfz, status_nfz = kod_alt, status_alt
            elif zameld["wojewodztwo"]:
                kod_region = dopasuj_nfz_region(zameld["wojewodztwo"])
                if kod_region:
                    kod_nfz, status_nfz = kod_region, "REGION"
        if status_nfz == "BRAK_SLOWNIKA":
            conflicts.append(("Slownik NFZ", "WYSOKI", "Brak pliku NFZ.csv w Slowniki"))
        elif status_nfz not in ("DOKLADNE", "ZAWIERA", "PRZYBLIZONE", "REGION"):
            conflicts.append(("NFZ", "WYSOKI", f"Nie dopasowano ({status_nfz}): {nfz_nazwa}"))
    data["Last.OddzialNFZ.Kod"] = kod_nfz

    return data, conflicts, firma_nazwa, firma_nip


# ======================= DUPLIKATY =======================

def compare_duplicates(rows_with_conflicts):
    """rows_with_conflicts: lista (data, conflicts, filename). Modyfikuje conflicts w miejscu."""
    by_pesel = {}
    for data, conflicts, filename in rows_with_conflicts:
        pesel = data.get("Last.PESEL", "")
        if not pesel:
            continue
        by_pesel.setdefault(pesel, []).append((data, conflicts, filename))

    duplikaty_plikow = set()
    for pesel, grupa in by_pesel.items():
        if len(grupa) < 2:
            continue
        pierwszy = grupa[0][0]
        pola_do_porownania = [c for c in ENOVA_COLUMNS if c not in ("NazwaPliku",)]
        rozne = False
        for data, _, _ in grupa[1:]:
            if any(data.get(c, "") != pierwszy.get(c, "") for c in pola_do_porownania):
                rozne = True
                break
        priorytet = "WYSOKI" if rozne else "NISKI"
        opis = "Ten sam PESEL w kilku plikach, dane sie roznia" if rozne else "Ten sam PESEL w kilku plikach, dane zgodne"
        for data, conflicts, filename in grupa:
            conflicts.append(("Duplikaty", priorytet, opis))
            duplikaty_plikow.add(filename)

    return duplikaty_plikow


# ======================= EXCEL =======================

def write_excel(rows, all_conflicts, path, duplikaty_plikow):
    wb = Workbook()
    ws = wb.active
    ws.title = "ENOVA_IMPORT"

    for col_idx, nazwa in enumerate(ENOVA_COLUMNS, start=1):
        c = ws.cell(row=1, column=col_idx, value=nazwa)
        c.font = Font(bold=True)

    idx_rachunek_kolumny = {
        ENOVA_COLUMNS.index("Rachunki:Rachunek.Numer.CS") + 1,
        ENOVA_COLUMNS.index("Rachunki:Rachunek.Numer.Kierunek") + 1,
        ENOVA_COLUMNS.index("Rachunki:Rachunek.Numer.Numer") + 1,
    }
    idx_plik = ENOVA_COLUMNS.index("NazwaPliku") + 1
    komorka_pliku = {}

    fill_duplikat = PatternFill(start_color=KOLOR_DUPLIKAT, end_color=KOLOR_DUPLIKAT, fill_type="solid")

    for r, data in enumerate(rows, start=2):
        for col_idx, nazwa in enumerate(ENOVA_COLUMNS, start=1):
            wartosc = data.get(nazwa, "")
            komorka = ws.cell(row=r, column=col_idx, value=wartosc)
            if col_idx in idx_rachunek_kolumny:
                komorka.number_format = "@"
            if data.get("NazwaPliku") in duplikaty_plikow:
                komorka.fill = fill_duplikat
        komorka_pliku[data.get("NazwaPliku")] = r

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col_idx, nazwa in enumerate(ENOVA_COLUMNS, start=1):
        max_len = max([len(str(nazwa))] + [len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(2, len(rows) + 2)])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

    if all_conflicts:
        ws2 = wb.create_sheet("KONFLIKTY")
        for col_idx, nazwa in enumerate(KONFLIKTY_COLUMNS, start=1):
            c = ws2.cell(row=1, column=col_idx, value=nazwa)
            c.font = Font(bold=True)

        kolumna_docelowa = {
            "Dane osobowe": "Last.Nazwisko", "PESEL": "Last.PESEL", "NIP": "Last.NIP",
            "Obywatelstwo": "Last.Obywatelstwo.Nazwa", "E-mail": "Kontakt.EMAIL",
            "Rachunek bankowy": "Rachunki:Rachunek.Numer.Numer", "Adres - Poczta": "Last.Adres.Poczta",
            "Urzad Skarbowy": "Last.Podatki.UrzadSkarbowy:Kod", "NFZ": "Last.OddzialNFZ.Kod",
            "Duplikaty": "Last.PESEL", "Rozne firmy": "NazwaPliku",
            "Przetwarzanie PDF": "NazwaPliku",
        }

        r = 2
        for filename, pesel, nazwisko, imie, priorytet, kategoria, problem in all_conflicts:
            ws2.cell(row=r, column=1, value=filename)
            ws2.cell(row=r, column=2, value=pesel)
            ws2.cell(row=r, column=3, value=nazwisko)
            ws2.cell(row=r, column=4, value=imie)
            ws2.cell(row=r, column=5, value=priorytet)
            ws2.cell(row=r, column=6, value=kategoria)
            komorka_problem = ws2.cell(row=r, column=7, value=problem)
            ws2.cell(row=r, column=8, value="Sprawdz i popraw w formularzu / arkuszu ENOVA_IMPORT")
            ws2.cell(row=r, column=9, value="")

            kol_docelowa = kolumna_docelowa.get(kategoria)
            wiersz_docelowy = komorka_pliku.get(filename)
            if kol_docelowa and wiersz_docelowy and kol_docelowa in ENOVA_COLUMNS:
                idx_kol = ENOVA_COLUMNS.index(kol_docelowa) + 1
                litera = get_column_letter(idx_kol)
                komorka_problem.hyperlink = f"#ENOVA_IMPORT!{litera}{wiersz_docelowy}"
                komorka_problem.font = Font(color="0563C1", underline="single")
            r += 1

        ws2.freeze_panes = "A2"
        ws2.auto_filter.ref = ws2.dimensions
        for col_idx in range(1, len(KONFLIKTY_COLUMNS) + 1):
            ws2.column_dimensions[get_column_letter(col_idx)].width = 22

    try:
        wb.save(path)
        return path
    except PermissionError:
        katalog = os.path.dirname(path)
        nazwa_bazowa = os.path.splitext(os.path.basename(path))[0]
        alternatywna = os.path.join(
            katalog, f"{nazwa_bazowa}_{datetime.now().strftime('%H_%M_%S')}.xlsx"
        )
        print(f"UWAGA: nie mozna zapisac {path} (plik jest prawdopodobnie otwarty w Excelu).")
        print(f"Zapisuje pod alternatywna nazwa: {alternatywna}")
        wb.save(alternatywna)
        return alternatywna


def verify_output_excel(path):
    if not os.path.isfile(path):
        print(f"BLAD: nie udalo sie zapisac pliku wynikowego {path}")
        return False
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True)
    if "ENOVA_IMPORT" not in wb.sheetnames:
        print("BLAD: brak arkusza ENOVA_IMPORT w pliku wynikowym")
        return False
    return True


# ======================= MAIN =======================

def main():
    print("odszyfruj_i_zestaw_enova_v16.py (v16_zabezpieczenie_pola_poczta)")
    print(f"Plik wynikowy: {PLIK_WYJSCIOWY_XLSX}")
    print(f"Folder na odszyfrowane PDF: {FOLDER_ODSZYFROWANE}")

    if not os.path.isdir(FOLDER_PDF):
        print("BLAD: nie istnieje folder z PDF-ami (PDF\\)")
        return

    pdf_files = sorted(glob.glob(os.path.join(FOLDER_PDF, "**", "*.pdf"), recursive=True))
    if not pdf_files:
        print("Brak plikow PDF w folderze PDF\\")
        return

    os.makedirs(FOLDER_ODSZYFROWANE, exist_ok=True)

    rows_with_conflicts = []  # (data, conflicts, filename)
    global_conflicts = []     # (filename, pesel, nazwisko, imie, priorytet, kategoria, problem)
    firmy = {}                # nazwa+nip -> lista plikow

    for path in pdf_files:
        filename = os.path.basename(path)
        try:
            reader = PdfReader(path)
            if reader.is_encrypted:
                wynik = reader.decrypt(HASLO)
                if wynik == 0:
                    print(f"Pominieto PDF zaszyfrowany innym haslem: {filename}")
                    continue
            else:
                print(f"PDF nie byl zaszyfrowany, przetwarzam mimo to: {filename}")

            writer = PdfWriter()
            for page in reader.pages:
                writer.add_page(page)
            out_path = os.path.join(FOLDER_ODSZYFROWANE, filename)
            with open(out_path, "wb") as f:
                writer.write(f)

            text = wyciagnij_tekst(reader, "default")

            data, conflicts, firma_nazwa, firma_nip = parse_pdf_text(text, filename)
            rows_with_conflicts.append((data, conflicts, data["NazwaPliku"]))

            if firma_nazwa or firma_nip:
                firmy.setdefault((firma_nazwa, firma_nip), []).append(filename)

            print(f"OK: {filename}")

        except Exception as e:
            global_conflicts.append((filename, "", "", "", "WYSOKI", "Przetwarzanie PDF", str(e)))
            print(f"BLAD przy {filename}: {e}")

    if not rows_with_conflicts:
        szczegoly_bledow = "\n".join(
            f"- {filename}: [{kategoria}] {problem}"
            for filename, _p, _n, _i, priorytet, kategoria, problem in global_conflicts
        )
        print("Brak poprawnie przetworzonych plikow.")
        if szczegoly_bledow:
            print("Szczegoly bledow dla poszczegolnych plikow:")
            print(szczegoly_bledow)
        pokaz_komunikat_windows(
            "Brak poprawnie przetworzonych plikow",
            "Zaden PDF nie zostal poprawnie przetworzony.\n\n"
            + (szczegoly_bledow if szczegoly_bledow else "Sprawdz folder PDF i haslo."),
            ikona="error",
        )
        return

    duplikaty_plikow = compare_duplicates(rows_with_conflicts)

    if len(firmy) > 1:
        lista_firm = ", ".join(f"{n or '(brak nazwy)'} (NIP {nip or 'brak'})" for n, nip in firmy)
        pokaz_komunikat_windows(
            "Rozne firmy w jednym imporcie",
            f"PDF-y pochodza z {len(firmy)} roznych firm:\n{lista_firm}\n\n"
            f"Enova importuje jedna firme na raz. Podziel pliki przed importem.",
            ikona="warning",
        )
        for (n, nip), pliki in firmy.items():
            for filename in pliki:
                global_conflicts.append((filename, "", "", "", "WYSOKI", "Rozne firmy",
                                          f"PDF z firmy {n or '(brak)'} NIP {nip or 'brak'}, "
                                          f"a w zestawie sa tez inne firmy"))

    rows = []
    for data, conflicts, filename in rows_with_conflicts:
        rows.append(data)
        for kategoria, priorytet, problem in conflicts:
            global_conflicts.append((
                filename, data.get("Last.PESEL", ""), data.get("Last.Nazwisko", ""),
                data.get("Last.Imie", ""), priorytet, kategoria, problem,
            ))

    faktyczna_sciezka = write_excel(rows, global_conflicts, PLIK_WYJSCIOWY_XLSX, duplikaty_plikow)

    if verify_output_excel(faktyczna_sciezka):
        print(f"\nZapisano zestawienie: {faktyczna_sciezka}")
        print(f"Przetworzono poprawnie: {len(rows)} plik(ow)")
        if global_conflicts:
            print(f"Konfliktow do sprawdzenia: {len(global_conflicts)} (arkusz KONFLIKTY)")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        import traceback
        szczegoly = traceback.format_exc()
        print(szczegoly)
        pokaz_komunikat_windows(
            "Skrypt zakonczyl sie bledem",
            f"{e}\n\nSzczegoly w konsoli / powyzej. Wklej ten komunikat do Claude.",
            ikona="error",
        )
